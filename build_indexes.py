"""Build all lookup indexes for the classification workspace."""
import openpyxl, json, os, re

BASE = "c:/Users/dgika/Downloads/Fw_ Υπενθύμιση_ Συμπληρωματικές Πληροφορίες (1)"
EXCEL = os.path.join(BASE, "excel")
WS = os.path.join(BASE, "_workspace")
DATA = os.path.join(WS, "data")
os.makedirs(DATA, exist_ok=True)

# ── 1. ACCOUNTS ──────────────────────────────────────────────────────────────
print("Building accounts index...")
wb = openpyxl.load_workbook(os.path.join(EXCEL, "1. ΛΟΓΙΣΤΙΚΟ ΣΧΕΔΙΟ.xlsx"), data_only=True)
accounts = {}
for sheet in wb.sheetnames:
    ws_sheet = wb[sheet]
    for row in ws_sheet.iter_rows(values_only=True):
        if not row or not row[1]: continue
        code = str(row[1]).strip()
        desc = str(row[2]).strip() if row[2] else ''
        if '.' not in code: continue
        parts = code.split('.')
        accounts[code] = {
            "desc": desc,
            "group": parts[0],
            "subgroup": parts[0]+"."+parts[1] if len(parts) > 1 else parts[0],
            "sheet": sheet
        }

with open(os.path.join(DATA, "accounts.json"), 'w', encoding='utf-8') as f:
    json.dump(accounts, f, ensure_ascii=False, indent=2)
print(f"  Saved {len(accounts)} accounts")

# ── 2. ACCOUNTS — reverse lookup: desc keywords → code ───────────────────────
# For each full-detail account (has 3 parts, len==11 or similar)
keyword_index = {}
for code, info in accounts.items():
    if code.count('.') < 2: continue  # skip group-level codes
    desc_upper = info['desc'].upper()
    # Extract significant keywords
    words = re.findall(r'[Α-ΩA-Z]{4,}', desc_upper)
    for w in words:
        if w not in keyword_index:
            keyword_index[w] = []
        keyword_index[w].append({"code": code, "desc": info['desc']})

with open(os.path.join(DATA, "accounts_by_keyword.json"), 'w', encoding='utf-8') as f:
    json.dump(keyword_index, f, ensure_ascii=False, indent=2)
print(f"  Saved keyword index with {len(keyword_index)} keywords")

# ── 3. ACCOUNTS — category × branch structure ─────────────────────────────────
# For the key variable-per-branch categories, build a subgroup → [codes] map
subgroup_index = {}
for code, info in accounts.items():
    sg = info['subgroup']
    if sg not in subgroup_index:
        subgroup_index[sg] = []
    subgroup_index[sg].append({"code": code, "desc": info['desc']})

with open(os.path.join(DATA, "accounts_by_subgroup.json"), 'w', encoding='utf-8') as f:
    json.dump(subgroup_index, f, ensure_ascii=False, indent=2)
print(f"  Saved subgroup index with {len(subgroup_index)} subgroups")

# ── 4. SUPPLIERS ──────────────────────────────────────────────────────────────
print("\nBuilding supplier indexes...")
wb2 = openpyxl.load_workbook(os.path.join(EXCEL, "6β. ΑΡΧΕΙΟ ΠΡΟΜΗΘΕΥΤΩΝ.xlsx"), data_only=True)
ws2 = wb2.active

by_afm = {}
by_code = {}
all_suppliers = []

header_found = False
for row in ws2.iter_rows(values_only=True):
    if not row or not row[0]: continue
    if str(row[0]).strip() == 'A/A':
        header_found = True
        continue
    if not header_found: continue
    if not isinstance(row[0], (int, float)): continue

    code = str(row[1]).strip() if row[1] else ''
    name = str(row[2]).strip() if row[2] else ''
    afm  = str(row[3]).strip() if row[3] else ''
    addr = str(row[4]).strip() if row[4] else ''
    area = str(row[5]).strip() if row[5] else ''
    city = str(row[7]).strip() if row[7] else ''

    if not code or not name: continue

    entry = {"code": code, "name": name, "afm": afm,
             "addr": addr, "area": area, "city": city}
    all_suppliers.append(entry)
    if afm and afm not in ('None', ''):
        by_afm[afm] = entry
    if code:
        by_code[code] = entry

with open(os.path.join(DATA, "suppliers_by_afm.json"), 'w', encoding='utf-8') as f:
    json.dump(by_afm, f, ensure_ascii=False, indent=2)
with open(os.path.join(DATA, "suppliers_by_code.json"), 'w', encoding='utf-8') as f:
    json.dump(by_code, f, ensure_ascii=False, indent=2)

print(f"  Saved {len(all_suppliers)} total suppliers")
print(f"  By AFM: {len(by_afm)}, By code: {len(by_code)}")

# ── 5. NAME SEARCH INDEX — first word of name → [suppliers] ──────────────────
name_index = {}
for s in all_suppliers:
    words = s['name'].upper().split()
    for w in words[:2]:  # index on first 2 words
        clean = re.sub(r'[^Α-ΩA-Z0-9]', '', w)
        if len(clean) >= 3:
            if clean not in name_index:
                name_index[clean] = []
            name_index[clean].append({"code": s['code'], "name": s['name'], "afm": s['afm']})

with open(os.path.join(DATA, "suppliers_by_name.json"), 'w', encoding='utf-8') as f:
    json.dump(name_index, f, ensure_ascii=False, indent=2)
print(f"  Name index: {len(name_index)} keywords")

print("\nAll indexes built. Data saved to: _workspace/data/")
